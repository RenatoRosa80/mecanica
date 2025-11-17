from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from financeiro.models import Lancamento
from clientes.models import Cliente
from estoque.models import Produto

# PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

# Excel
import openpyxl
from openpyxl.styles import Font
from io import BytesIO

from datetime import datetime


# ========================
# PÁGINA INICIAL DOS RELATÓRIOS
# ========================

def relatorios_home(request):
    return render(request, 'relatorios/relatorios_home.html')


# ========================
# RELATÓRIO FINANCEIRO (HTML)
# ========================

def relatorio_financeiro(request):
    data_inicio = request.GET.get('inicio')
    data_fim = request.GET.get('fim')

    registros = Lancamento.objects.all().order_by('-data')

    if data_inicio and data_fim:
        registros = registros.filter(data__range=[data_inicio, data_fim])

    total_entradas = registros.filter(tipo='E').aggregate(Sum('valor'))['valor__sum'] or 0
    total_saidas = registros.filter(tipo='S').aggregate(Sum('valor'))['valor__sum'] or 0
    saldo = total_entradas - total_saidas

    context = {
        'registros': registros,
        'total_entradas': total_entradas,
        'total_saidas': total_saidas,
        'saldo': saldo,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }

    return render(request, 'relatorios/relatorio_financeiro.html', context)


# ========================
# RELATÓRIO FINANCEIRO (PDF)
# ========================

def relatorio_financeiro_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    y = 800

    p.setFont("Helvetica-Bold", 16)
    p.drawString(170, y, "Relatório Financeiro")
    y -= 40

    registros = Lancamento.objects.all().order_by('-data')

    p.setFont("Helvetica", 12)
    for item in registros:
        tipo = "Entrada" if item.tipo == "E" else "Saída"
        linha = f"{item.data} - {item.descricao} - {tipo} - R$ {item.valor}"
        p.drawString(50, y, linha)
        y -= 20

        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 12)
            y = 800

    p.save()
    return response


# ========================
# RELATÓRIO FINANCEIRO (EXCEL)
# ========================

def relatorio_financeiro_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financeiro"

    ws.append(["Data", "Descrição", "Tipo", "Valor"])

    bold = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold

    registros = Lancamento.objects.all()

    for r in registros:
        ws.append([
            r.data.strftime("%d/%m/%Y"),
            r.descricao,
            "Entrada" if r.tipo == "E" else "Saída",
            float(r.valor)
        ])

    # CORREÇÃO: Substituição do save_virtual_workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.xlsx"'
    return response


# ========================
# RELATÓRIO CLIENTES
# ========================

def relatorio_clientes(request):
    clientes = Cliente.objects.all().order_by('nome')
    return render(request, 'relatorios/relatorio_clientes.html', {'clientes': clientes})


def relatorio_clientes_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_clientes.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    y = 800

    p.setFont("Helvetica-Bold", 16)
    p.drawString(170, y, "Relatório de Clientes")
    y -= 40

    clientes = Cliente.objects.all().order_by("nome")

    p.setFont("Helvetica", 12)
    for c in clientes:
        linha = f"{c.nome} - CPF: {c.cpf} - Tel: {c.telefone}"
        p.drawString(50, y, linha)
        y -= 20

        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 12)
            y = 800

    p.save()
    return response


def relatorio_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    ws.append(["Nome", "CPF", "Telefone", "Email", "Endereço"])

    bold = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold

    clientes = Cliente.objects.all()

    for c in clientes:
        ws.append([
            c.nome,
            c.cpf,
            c.telefone,
            c.email or "",
            c.endereco or ""
        ])

    # CORREÇÃO: Substituição do save_virtual_workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename="relatorio_clientes.xlsx"'
    return response


# ========================
# RELATÓRIO ESTOQUE
# ========================

def relatorio_estoque(request):
    produtos = Produto.objects.all()
    
    # Calcular o valor total para cada produto
    produtos_com_total = []
    for produto in produtos:
        valor_total = produto.quantidade * produto.preco_unitario
        produtos_com_total.append({
            'produto': produto,
            'valor_total': valor_total
        })
    
    return render(request, 'relatorios/relatorio_estoque.html', {
        'produtos_com_total': produtos_com_total
    })


def relatorio_estoque_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    y = 800

    p.setFont("Helvetica-Bold", 16)
    p.drawString(170, y, "Relatório de Estoque")
    y -= 40

    produtos = Produto.objects.all()

    p.setFont("Helvetica", 12)
    for prod in produtos:
        linha = f"{prod.nome} - Qtde: {prod.quantidade} - R$ {prod.preco_unitario}"
        p.drawString(50, y, linha)
        y -= 20

        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 12)
            y = 800

    p.save()
    return response


def relatorio_estoque_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"

    ws.append(["Nome", "Quantidade", "Preço Unitário", "Valor Total"])

    bold = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold

    produtos = Produto.objects.all()

    for prod in produtos:
        valor_total = float(prod.quantidade * prod.preco_unitario)
        ws.append([
            prod.nome,
            prod.quantidade,
            float(prod.preco_unitario),
            valor_total
        ])

    # CORREÇÃO: Substituição do save_virtual_workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.xlsx"'
    return response


# ========================
# RELATÓRIO COMPLETO (TODOS OS DADOS)
# ========================

def relatorio_completo_excel(request):
    wb = openpyxl.Workbook()
    
    # ABA FINANCEIRO
    ws_fin = wb.active
    ws_fin.title = "Financeiro"
    ws_fin.append(["Data", "Descrição", "Tipo", "Valor"])
    
    bold = Font(bold=True)
    for cell in ws_fin["1:1"]:
        cell.font = bold
    
    registros = Lancamento.objects.all()
    for r in registros:
        ws_fin.append([
            r.data.strftime("%d/%m/%Y"),
            r.descricao,
            "Entrada" if r.tipo == "E" else "Saída",
            float(r.valor)
        ])
    
    # ABA CLIENTES
    ws_cli = wb.create_sheet("Clientes")
    ws_cli.append(["Nome", "CPF", "Telefone", "Email", "Endereço"])
    
    for cell in ws_cli["1:1"]:
        cell.font = bold
    
    clientes = Cliente.objects.all()
    for c in clientes:
        ws_cli.append([
            c.nome,
            c.cpf,
            c.telefone,
            c.email or "",
            c.endereco or ""
        ])
    
    # ABA ESTOQUE
    ws_est = wb.create_sheet("Estoque")
    ws_est.append(["Nome", "Quantidade", "Preço Unitário", "Valor Total"])
    
    for cell in ws_est["1:1"]:
        cell.font = bold
    
    produtos = Produto.objects.all()
    for prod in produtos:
        valor_total = float(prod.quantidade * prod.preco_unitario)
        ws_est.append([
            prod.nome,
            prod.quantidade,
            float(prod.preco_unitario),
            valor_total
        ])

    # CORREÇÃO: Substituição do save_virtual_workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename="relatorio_completo.xlsx"'
    return response